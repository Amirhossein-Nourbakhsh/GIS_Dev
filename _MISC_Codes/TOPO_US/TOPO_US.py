#-------------------------------------------------------------------------------
# Name:        TOPO_US
# Purpose:     Converts USGS current Topographic maps from geoPDF to geoTIF using GDAL. 
#
# Author:      cchen
#
# Created:     16/08/2016
# Copyright:   (c) cchen 2016
# Licence:     <your licence>
# Updated:      2021-01-13
#-------------------------------------------------------------------------------
'''2020-11-27 AW - I believe the previous script used the NEATLINE from GDALINFO.exe to crop the image from geoPDFs,
but the neatline from this new batch of topo geopdfs crops to the edge of the pdf instead of the image, so I used
the bounding coordinates from the XML metafile.
The bounding coordinates from the XML are in WGS84, so I converted the coordinates from WGS84 to the geoPDF's assigned
coordinate system to create the neatline. However, this caused the image to not crop properly. Due to this, I don't think
converting the bounding coordinates is necessary. Just use the original bounding coordinates in WGS84 format as the neatline,
but when running the GDALWARP.exe (cropping process), set the target coodinate system t_srs as WGS84 (4326).
The output image .tif will looked skewed in photoviewer, but will look fine in Arcmap when viewed in the geoPDF's coordinate system.
    Using:
    Python 3.7.9
    GDAL 3.2.1'''

import multiprocessing as mp
import multiprocessing.pool as mpp
import os, glob, sys, logging, time
import traceback
import openpyxl
import csv
import TOPO_US_single as TOPO_single

class NoDaemonProcess(mp.Process):
    def _get_daemon(self):
        return False
    def _set_daemon(self, value):
        pass
    daemon = property(_get_daemon, _set_daemon)

class MyPool(mpp.Pool):
    Process = NoDaemonProcess

class countlist():
    def countcurrenttif(self,tifdirExist):
        for file in os.listdir(tifdirExist):
            if file.endswith(".tif"):
                tifdirlist.append(file)
        print(f"number of exist topo tifs: {len(tifdirlist)}")
        return tifdirlist

    def countnewpdf(self,pdfdirNew):
        for file in os.listdir(pdfdirNew):
            if file.endswith(".pdf"):
                pdfdirlist.append(file)
        print(f"number of new topo pdfs: {len(pdfdirlist)}")
        return pdfdirlist

    def checkfile(self,pdfdirlist,tifdirlist):
        for pdfname in pdfdirlist:
            name = pdfname[0:-4]
            if name+"_t.tif" in tifdirlist and name+"_o.tif" in tifdirlist:
                donelist.append(pdfname)
            else:
                newlist.append(pdfname)
        print(f"number of pdfs that exists already: {len(donelist)}")
        print(f"number of pdfs to convert: {len(newlist)}")
        return donelist, newlist

    def writecount(self, mscpath):
        with open(fr'{mscpath}\topo_existTopo.txt','w') as txtfile:
            txtfile.writelines("%s\n" % name for name in tifdirlist)
        with open(fr'{mscpath}\topo_doneTopo.txt','w') as txtfile:
            txtfile.writelines("%s\n" % name for name in donelist)
        with open(fr'{mscpath}\topo_newTopo.txt','w') as txtfile1:
            txtfile1.writelines( "%s\n" % name for name in newlist)

class work():
    def getjoinpaths(self,pdfname):
        pdfpath = os.path.join(pdfdirNew,pdfname)
        tifpath_t = os.path.join(tifdirNew,pdfname[0:-4]+"_t.tif")
        tifpath_o = os.path.join(tifdirNew,pdfname[0:-4]+"_o.tif")
        xmlpath = os.path.join(tifdirNew,pdfname[0:-4]+".xml")
        csvtemppath = os.path.join(mscpath, pdfname[0:-4]+'.csv')
        logpath = os.path.join(mscpath, 'TOPO_US_log.txt')
        return pdfpath,tifpath_t,tifpath_o,xmlpath,csvtemppath,logpath

    def poolhandler(self,list):
        # # SINGLE THREAD
        # n = 1
        # for pdf in list:
        #     print("------------------------------------------------------")
        #     print(n)            
        #     t = self.getjoinpaths(pdf) + (gdalpath, xmldict, pdf,)
        #     self.poolworker(t)
        #     n += 1
        # print(f"...completed {n-1} records.")
        # return xmldict
        # # RI took 184.85 mins

        # MULTIPROCESSING POOL
        modlist = []
        for pdf in list:
            t = self.getjoinpaths(pdf) + (gdalpath, xmldict, pdf,)
            modlist.append(t)
            
        p = MyPool(processes)

        for x in p.map(self.poolworker,modlist):                      # RUN WORKER AND RETRUN DICT (STILL IN LIST FORM)                             
            xmldict.update(x)                                         # CONVERT LIST TO BACK DICT

        # for x in p.map_async(self.poolworker,modlist).get():
        #     xmldict.update(x)
        return xmldict
        # RI took 51.03 mins - 60.38 mins

    def poolworker(self,items):
        tp = TOPO_single.TOPO_Process()
        pdfpath, tifpath_t, tifpath_o, xmlpath, csvtemppath, logpath, gdalpath, xmldict, pdfname = items
        tp.main(pdfname, pdfpath, tifpath_t, tifpath_o, xmlpath, csvtemppath, gdalpath, xmldict)
        self.createlog(items)
        return xmldict

    def createlog(self,items):
        pdfpath, tifpath_t, tifpath_o, xmlpath, csvtemppath, logpath, gdalpath, xmldict, pdfname = items

        # CREATE LOG
        handler = logging.FileHandler(logpath)
        logger = logging.getLogger('totiff')
        logger.setLevel(logging.INFO)        
        logger.addHandler(handler)

        if not os.path.exists(xmlpath) or not os.path.exists(tifpath_t) or os.path.getsize(tifpath_t)  < 1000000 or not os.path.exists(tifpath_o) or os.path.getsize(tifpath_o)  < 1000000:
            logger.info("------------------------------------------------------")
            logger.info(time.strftime('%Y-%m-%d %H:%M:%S', time.localtime()))
            logger.info(pdfname)

            # CHECK XML EXIST
            if not os.path.exists(xmlpath):
                logger.error("### Cannot find .xml file...")

            # CHECK IMAGE_t
            if not os.path.exists(tifpath_t):
                logger.error("### Cannot find .tif file...")
            elif os.path.getsize(tifpath_t)  < 1000000:   #check if file size is less than 1M
                logger.error("Error (black image): " + tifpath_t + " (problem with writing file).")

            # CHECK IMAGE_o
            if not os.path.exists(tifpath_o):
                logger.error("### Cannot find .tif file...")
            elif os.path.getsize(tifpath_o)  < 1000000:   #check if file size is less than 1M
                logger.error("Error (black image): " + tifpath_o + " (problem with writing file).")

        logger.debug(pdfname + "  " + time.strftime('%Y-%m-%d %H:%M:%S', time.localtime()))
        logger.removeHandler(handler)

        # remove temp csv $ empty log
        if os.path.getsize(logpath) == 0:
            os.remove(logpath)

def createxlsx(metadict, xlsxpath, csvpath):
    print("--------------------")
    print("...write to xlsx...")

    # MERGE XML AND CSV TO DICT
    with open(csvpath, "r") as f:
        reader = csv.reader(f)
        for row in reader:
            key = row[15]
            if key in metadict:
                row.reverse()
                for item in row:
                    metadict[key].insert(0, item)

    # WRITE TO XLSX
    if os.path.exists(xlsxpath):
        wb = openpyxl.load_workbook(xlsxpath)
        ws = wb["TOPO"]
        
        r = ws.max_row + 1
        print(r)
        for k, v in metadict.items():
            for i in range(0, 33):
                ws.cell(row=r, column=(i+1), value=v[i]).number_format = "@"
            ws.cell(row=r, column=34, value=record_date).number_format = "@"
            r += 1
    else:
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "TOPO"
        
        # ADD XLSX HEADERS  *openpyxl starts at 1
        ws.cell(row=1, column=1).value = "Item ID"
        ws.cell(row=1, column=2).value = "Download(MB unzipped)"
        ws.cell(row=1, column=3).value = "BA"
        ws.cell(row=1, column=4).value = "Sa"
        ws.cell(row=1, column=5).value = "Se"
        ws.cell(row=1, column=6).value = "Grid Size"
        ws.cell(row=1, column=7).value = "Short Name"
        ws.cell(row=1, column=8).value = "Full Name"
        ws.cell(row=1, column=9).value = "MassLoad ID"
        ws.cell(row=1, column=10).value = "Cell ID"
        ws.cell(row=1, column=11).value = "Create Date"
        ws.cell(row=1, column=12).value = "SourceYear"
        ws.cell(row=1, column=13).value = "HTMCImprintYear"
        ws.cell(row=1, column=14).value = "HTMCScan ID"
        ws.cell(row=1, column=15).value = "Download URL"
        ws.cell(row=1, column=16).value = "Filename"
        ws.cell(row=1, column=17).value = "Cloud URL"

        ws.cell(row=1, column=18).value = "TOPO_FLAG"
        ws.cell(row=1, column=19).value = "T_TFW_FLAG"
        ws.cell(row=1, column=20).value = "ORTHO_FLAG"
        ws.cell(row=1, column=21).value = "O_TFW_FLAG"
        ws.cell(row=1, column=22).value = "XML_FLAG"
        ws.cell(row=1, column=23).value = "EPSG"
        ws.cell(row=1, column=24).value = "WKT"
        ws.cell(row=1, column=25).value = "NEATLINE"

        ws.cell(row=1, column=26).value = "XML_DATE_ON_MAP"
        ws.cell(row=1, column=27).value = "XML_IMPRINT_YEAR"
        ws.cell(row=1, column=28).value = "XML_PHOTO_INSPECTION_YEAR"
        ws.cell(row=1, column=29).value = "XML_PHOTO_REVISION_YEAR"
        ws.cell(row=1, column=30).value = "XML_FIELD_CHECK_YEAR"
        ws.cell(row=1, column=31).value = "XML_SURVEY_YEAR"
        ws.cell(row=1, column=32).value = "XML_EDIT_YEAR"
        ws.cell(row=1, column=33).value = "XML_AERIAL_PHOTO_YEAR"

        ws.cell(row=1, column=34).value = "ERIS_RECORD_DATE"

        # WRITE DICT TO XLSX
        r = 2
        for k, v in metadict.items():
            for i in range(0, 33):
                ws.cell(row=r, column=(i+1), value=v[i]).number_format = "@"
            ws.cell(row=r, column=34, value=record_date).number_format = "@"
            r += 1
    wb.save(xlsxpath)
    wb.close

# MAIN ==================================================================================================================================================
if __name__ == '__main__':
    start = time.perf_counter()

    # PATHS TO CHANGE
    pdfdirNew =r"F:\A Wong_USTopo_10-7-20\USGS_US TOPO Dataset_9-18-20"                         # PATH TO NEW TOPO PDF DATA/HARD DRIVE
    tifdirNew = r"\\cabcvan1fpr009\USGS_Topo\USGS_currentTopo_Geotiff\new_current"              # PATH TO STORE NEW CONVERTED TOPO TIFS (EMPTY FOLDER)
    tifdirExist = r"\\cabcvan1fpr009\USGS_Topo\USGS_currentTopo_Geotiff"                        # PATH TO CURRENT EXISTING CONVERTED TOPO TIFS
    mscpath = r"\\cabcvan1gis005\MISC_DataManagement\Data\US\_FED\TOPO\2020_09_18"              # PATH TO STORE LOGS/EXCEL
    gdalpath = r"C:\Program Files\GDAL"                                                         # PATH TO GDAL
    
    csvfile = "USTopo_all_T_7.5_gda_results.csv"                                                # NEW CSV FILE NAME
    record_date = "2020-09-18"                                                                  # RECORD DATE OF NEW DATA
    state = ["FL"]                                                                              # STATE FILTER
    processes = 12                                                                           # INPUT NUMBER OF CORES/PROCESSES TO RUN ACCORDING TO YOUR PC SPECS (os/mp.cpu_count()). PLEASE MONITOR YOUR CPU TEMPERATURE.
    # ----------------------------------------------------------------------------------
    csvpath = os.path.join(mscpath, "_RAW", csvfile)
    xlsxpath = os.path.join(mscpath, "TOPO_US.xlsx")

    # Freeze Support for Multiprocessing
    mp.freeze_support()

    ct = countlist()
    wk = work()
    pdfdirlist=[]
    tifdirlist=[]
    newlist=[]
    donelist=[]
    xmldict = {} 
    # ----------------------------------------------------------------------------------
    
    print(time.strftime('%Y-%m-%d %H:%M:%S', time.localtime()))
    # CHECKS AND CREATE LISTS
    if not os.path.exists(os.path.join(mscpath, "topo_newTopo.txt")):
        # COUNT CURRENT EXIST NUMBER OF TOPO TIFS IN INVENTORY
        ct.countcurrenttif(tifdirExist)
    
        # COUNT NEW NUMBER OF TOPO PDFS RECEIVED
        ct.countnewpdf(pdfdirNew)

        # CHECK WHICH PDFS NEEDS TO BE CONVERTED(NEWLIST) AND WHICH CAN BE SKIPPED(DONELIST)
        ct.checkfile(pdfdirlist,tifdirlist)

        # WRITE COUNT
        ct.writecount(mscpath)
    else:
        # SKIP RUNNING PREVIOUS LINES OF CHECKS AND CREATING LISTS AGAIN
        with open(os.path.join(mscpath, "topo_newTopo.txt")) as f:
            newlist = f.read().splitlines()
            f.close()

    # CONVERT PDFS TO TIFS    
    if state:
        newlist = [p for p in newlist if p.split("_")[0] in state]
        wk.poolhandler(newlist)         # RUN POOL HANDLER AND RETURN XMLDICT
    else:      
        wk.poolhandler(newlist)         # RUN POOL HANDLER AND RETURN XMLDICT

    # CREATE XLSX
    createxlsx(xmldict, xlsxpath, csvpath)
    
    finish = time.perf_counter()
    print("--------------------")
    print(len(newlist))
    print(f"Process finished in {round((finish-start)/60, 2)} mins  --  {time.strftime('%Y-%m-%d %H:%M:%S', time.localtime())}")