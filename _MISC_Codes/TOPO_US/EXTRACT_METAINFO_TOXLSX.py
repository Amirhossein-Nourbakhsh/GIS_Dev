#-------------------------------------------------------------------------------
# Name:        TOPO_US-all
# Purpose:     As we did not have a compiled list of all our topo inventory combined, this script will
#              extract info from the csv (merged csv of all old csvs) and xmls from our current inventory, and compile it into one list.
# Date:        20210413
#-------------------------------------------------------------------------------

import xml.etree.ElementTree as ET
import os
import re
import time
import csv
import openpyxl
import datetime
import multiprocessing as mp
import multiprocessing.pool as mpp

class NoDaemonProcess(mp.Process):
    def _get_daemon(self):
        return False
    def _set_daemon(self, value):
        pass
    daemon = property(_get_daemon, _set_daemon)

class MyPool(mpp.Pool):
    Process = NoDaemonProcess

class TOPO_Process():

    def poolworker(self, file, tifdir, xmldict):
        pdfname = file[0:-6]+".pdf"
        tifpath_t = os.path.join(tifdir, pdfname[0:-4]+"_t.tif")
        tifpath_o = os.path.join(tifdir, pdfname[0:-4]+"_o.tif")
        xmlpath = os.path.join(tifdir,pdfname[0:-4]+".xml")
        self.getmetadata(pdfname, tifpath_t, tifpath_o, xmlpath, xmldict)
        return xmldict

    def getmetadata(self, pdfname, tifpath_t, tifpath_o, xmlpath, xmldict):     
        if pdfname == "" or pdfname == None or pdfname == " ":
            print("there's a blank!!!")

        # GET BOUNDING COORDINATES/NEATLINE FROM XML FILE
        westbc = None
        eastbc = None
        northbc = None
        southbc = None
    
        yeardateonmap = None
        yearimprint = None
        yearphotoinsp = None
        yearphotorevis = None
        yearfieldcheck = None
        yearsurvey = None
        yearedit = None
        yearaerial = None
        neatline_wkt = None
    
        if os.path.exists(xmlpath):
            tree = ET.parse(xmlpath)
            root = tree.getroot()
        
            # procsteps = root.findall("./dataqual/lineage/procstep")
            # yeardict = {}
            # for procstep in procsteps:
            #     procdate = procstep.find("./procdate")
            #     if procdate != None:
            #         procdesc = procstep.find("./procdesc")
            #         yeardict[procdesc.text.upper()] = procdate.text

            for child in root.iter():
                if child.tag == 'westbc':
                    westbc = float(child.text)
                elif child.tag == 'eastbc':
                    eastbc = float(child.text)
                elif child.tag == 'northbc':
                    northbc = float(child.text)
                elif child.tag == 'southbc':
                    southbc = float(child.text)
                elif child.tag == 'procstep':
                    if child[0].text.upper().strip() == 'DATE ON MAP':
                        yeardateonmap = child[1].text.strip()
                    elif child[0].text.upper().strip() == 'IMPRINT YEAR':
                        yearimprint = child[1].text.strip()
                    elif child[0].text.upper().strip() == 'PHOTO INSPECTION YEAR':
                        yearphotoinsp = child[1].text.strip()
                    elif child[0].text.upper().strip() == 'PHOTO REVISION YEAR':
                        yearphotorevis = child[1].text.strip()
                    elif child[0].text.upper().strip() == 'FIELD CHECK YEAR':
                        yearfieldcheck = child[1].text.strip()                                                                                      
                    elif child[0].text.upper().strip() == 'SURVEY YEAR':
                        yearsurvey = child[1].text.strip()
                    elif child[0].text.upper().strip() == 'EDIT YEAR':
                        yearedit = child[1].text.strip()
                    elif child[0].text.upper().strip() == 'AERIAL PHOTO YEAR':
                        yearaerial = child[1].text.strip()     
        
            neatline_wkt = '"' + f"POLYGON (({eastbc} {southbc}, {eastbc} {northbc}, {westbc} {northbc}, {westbc} {southbc}, {eastbc} {southbc}))" + '"'
    
        # CREATE DICT FOR USE IN METAXLSX FILE
        if os.path.exists(xmlpath):
            xmlFLAG = "Y"
        else:
            xmlFLAG = ""            
        if os.path.exists(tifpath_t):
            topoFLAG = "Y"
        else:
            topoFLAG = ""
        if os.path.exists(tifpath_t[0:-4]+".tfw"):
            t_tfwFLAG = "Y"
        else:
            t_tfwFLAG = ""
        if os.path.exists(tifpath_o):
            orthoFLAG = "Y"
        else:
            orthoFLAG = ""
        if os.path.exists(tifpath_o[0:-4]+".tfw"):
            o_tfwFLAG = "Y"
        else:
            o_tfwFLAG = ""
        
        tgtepsg = ""
        wkt = ""

        xmldict[pdfname] = [topoFLAG, t_tfwFLAG, orthoFLAG, o_tfwFLAG, xmlFLAG, tgtepsg, wkt, neatline_wkt, yeardateonmap, yearimprint, yearphotoinsp, yearphotorevis, yearfieldcheck, yearsurvey, yearedit, yearaerial]
        return xmldict

    def mergemetadata(self,metadict,csvpath):
        print("--------------------")
        print("...merge xml to xlsx...")

        # MERGE XML AND CSV TO DICT
        with open(csvpath, "r") as f:
            reader = csv.DictReader(f)
            for row in reader:
                try:
                    key = row["Filename"]
                    if key in ["", " ", None] or ".pdf" not in key:
                        continue

                    line = [row["Item ID"], row["Download(MB unzipped)"], row["BA"], row["Sa"], row["Se"], row["Grid Size"], row["Short Name"], row["Full Name"], row["MassLoad ID"], row["Cell ID"], row["Create Date"], row["SourceYear"], row["HTMCImprintYear"], row["HTMCScan ID"], row["Download URL"], row["Filename"], ""]
                    csvdate = datetime.datetime.strptime(row["Create Date"], "%d-%b-%y")

                    if key not in metadict:
                        metadict[key] = line
                    elif key in metadict:
                        if len(metadict[key]) == 16:
                            newline = line + metadict[key]
                            del metadict[key]
                            metadict[key] = newline
                        elif len(metadict[key]) == 17:                                  #  use the most recent record
                            metadate = datetime.datetime.strptime(metadict[key][10], "%d-%b-%y")
                            if metadate < csvdate:
                                del metadict[key]
                                metadict[key] = newline
                        elif len(metadict[key]) == 33:                                  #  use the most recent record
                            metadate = datetime.datetime.strptime(metadict[key][10], "%d-%b-%y")
                            if metadate < csvdate:
                                del metadict[key][:17]
                                newline = line + metadict[key]
                                del metadict[key]
                                metadict[key] = newline
                        else:
                            print(row)
                            print(metadict[key])
                            print(len(row),len(metadict[key]))
                            input(">>>")
                    else:
                        print(row)
                except:
                    print(row)
                    raise

    def createxlsx(self, metadict, outxlsx, csvpath):
        # WRITE TO XLSX
        if os.path.exists(outxlsx):
            wb = openpyxl.load_workbook(outxlsx)
            ws = wb[topotype]
            
            r = ws.max_row + 1
            for k, v in metadict.items():
                for i in range(0, 33):
                    ws.cell(row=r, column=(i+1), value=v[i]).number_format = "@"
                ws.cell(row=r, column=34, value=record_date).number_format = "@"
                r += 1
        else:
            wb = openpyxl.Workbook()
            ws = wb.active
            ws.title = topotype
            
            # ADD XLSX HEADERS  *openpyxl starts at 1
            ws.cell(row=1, column=1).value = "Item ID"
            ws.cell(row=1, column=2).value = "Download(MB unzipped)"
            ws.cell(row=1, column=3).value = "BA"
            ws.cell(row=1, column=4).value = "Sa"
            ws.cell(row=1, column=5).value = "Se"
            ws.cell(row=1, column=6).value = "Grid Size"
            ws.cell(row=1, column=7).value = "Short Name"
            ws.cell(row=1, column=8).value = "Full Name"
            ws.cell(row=1, column=9).value = "MassLoad ID"
            ws.cell(row=1, column=10).value = "Cell ID"
            ws.cell(row=1, column=11).value = "Create Date"
            ws.cell(row=1, column=12).value = "SourceYear"
            ws.cell(row=1, column=13).value = "HTMCImprintYear"
            ws.cell(row=1, column=14).value = "HTMCScan ID"
            ws.cell(row=1, column=15).value = "Download URL"
            ws.cell(row=1, column=16).value = "Filename"
            ws.cell(row=1, column=17).value = "Cloud URL"

            ws.cell(row=1, column=18).value = "TOPO_FLAG"
            ws.cell(row=1, column=19).value = "T_TFW_FLAG"
            ws.cell(row=1, column=20).value = "ORTHO_FLAG"
            ws.cell(row=1, column=21).value = "O_TFW_FLAG"
            ws.cell(row=1, column=22).value = "XML_FLAG"
            ws.cell(row=1, column=23).value = "EPSG"
            ws.cell(row=1, column=24).value = "WKT"
            ws.cell(row=1, column=25).value = "NEATLINE"

            ws.cell(row=1, column=26).value = "XML_DATE_ON_MAP"
            ws.cell(row=1, column=27).value = "XML_IMPRINT_YEAR"
            ws.cell(row=1, column=28).value = "XML_PHOTO_INSPECTION_YEAR"
            ws.cell(row=1, column=29).value = "XML_PHOTO_REVISION_YEAR"
            ws.cell(row=1, column=30).value = "XML_FIELD_CHECK_YEAR"
            ws.cell(row=1, column=31).value = "XML_SURVEY_YEAR"
            ws.cell(row=1, column=32).value = "XML_EDIT_YEAR"
            ws.cell(row=1, column=33).value = "XML_AERIAL_PHOTO_YEAR"

            ws.cell(row=1, column=34).value = "ERIS_RECORD_DATE"

            # WRITE DICT TO XLSX
            n = 0
            r = 2
            for k, v in metadict.items():
                if v[15] not in ["", " ", None] and ".pdf" in v[15]:
                    try:
                        if len(v) == 17:
                            for i in range(0,17):
                                ws.cell(row=r, column=(i+1), value=v[i]).number_format = "@"
                        elif len(v) == 33:
                            for i in range(0, 33):
                                ws.cell(row=r, column=(i+1), value=v[i]).number_format = "@"
                        else:
                            print(k,v)
                        ws.cell(row=r, column=34, value=record_date).number_format = "@"
                        r += 1
                    except Exception as e:
                        print(f"...failed...{i, k, v}")
                        raise
                else:
                    print("there's a blank!")
                    print(k,v)
                    n+=1
                    continue
            print(f"\t{n} csv records not in metadict..." )

        wb.save(outxlsx)
        wb.close

# ===============================================================================================================
if __name__ == '__main__':
    s = time.perf_counter()
    mp.freeze_support()
    tifdir = r"\\CABCVAN1FPR009\USGS_Topo\USGS_HTMC_Geotiff"
    mscpath = r"\\cabcvan1fpr001\data_us\_GIS_Layers\_FEDERAL\TOPO\2020_09_18"
    csvfile = r"\\cabcvan1fpr001\data_us\_GIS_Layers\_FEDERAL\TOPO\2020_09_18\_oldcsv\HTMC_US_merged.csv"
    record_date = ""
    topotype = "HTMC"

    outxlsx = os.path.join(mscpath, f"{topotype}_US_xmlcsv.xlsx")
    xmldict = {}
    tp = TOPO_Process()
    processes = 100

    print(f"...starting {time.strftime('%Y-%m-%d %H:%M:%S', time.localtime())}")
    tiflist = [(file, tifdir, xmldict) for file in os.listdir(tifdir) if file.endswith("_t.tif")]
    p = MyPool(processes)
    for x in p.starmap(tp.poolworker, tiflist):                      # RUN WORKER AND RETURN DICT (STILL IN LIST FORM)                             
        xmldict.update(x)    
    
    print(len(xmldict))
    tp.mergemetadata(xmldict,csvfile)
    tp.createxlsx(xmldict,outxlsx,csvfile)

    f = time.perf_counter()
    print(f"finished in {round(f-s, 2)} secs  --  {time.strftime('%Y-%m-%d %H:%M:%S', time.localtime())}")